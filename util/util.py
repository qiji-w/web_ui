import csv
import os
import sys
import time
import traceback
import random
import string
import zipfile
from configparser import ConfigParser
from urllib.parse import urlparse

import yagmail as yagmail
from dingtalkchatbot.chatbot import DingtalkChatbot
import requests
import yaml
import records
import pymysql
import openpyxl
from requests.cookies import RequestsCookieJar

from conf.conf import MYSQL_PATH
from util.log import Logger
import allure

project_path = os.path.dirname(os.path.dirname(__file__))
log_base_dir = os.path.join(project_path, 'log')
log_file = os.path.join(log_base_dir, 'all.log')
logging = Logger(log_file).logger

_error_num = 1


# 生成随机字符串
def generate_random_string(length: int = 8, prefix: str = ''):
    return f"{prefix}{'_' if prefix else ''}{''.join(random.choices(string.ascii_uppercase + string.digits, k=length))}"


# 获取元素 报错截图装饰器
def element_add_screen_and_log(func):
    def wrapper(*args, **kwargs):
        from basepage.base_page import BasePage
        _max_num = 3
        global _error_num
        # 获取实例化的对象
        instance: BasePage = args[0]
        try:
            logging.info("run " + func.__name__ + "\n args: \n" + repr(args[1:]) + "\n" + repr(kwargs))
            element = func(*args, **kwargs)
            _error_num = 1

            # 复原来的隐式等待时间
            instance.set_implicitly_wait(5)
            return element

        except Exception as e:
            # instance.screen_shot(inspect.stack()[2].function)
            logging.error("element operation failed".center(56, '*'))
            # 重置隐式等待时间
            # instance.set_implicitly_wait(1)

            # 判断异常处理次数
            if _error_num >= _max_num:
                logging.info("=" * 60)
                browser_name = instance._driver.desired_capabilities.get('browserName')
                #  chrome console日志写入日志
                if browser_name == 'chrome':
                    console_log = instance._driver.get_log('browser')
                    logging.info(" console log start ".center(36, '*'))
                    # for entry in console_log:
                    # logging.error(f'level: {entry.get("level")}, message: {entry.get("message")}')
                    logging.error(console_log)
                    logging.info(" console log end ".center(36, '*'))
                logging.error(f"current url is {instance._driver.current_url}")

                # 将截图插入allure报告中
                allure.attach(body=instance.screen_shot(), attachment_type=allure.attachment_type.PNG)
                print("******************************************************")
                logging.error(e)
                logging.info("=" * 60)

                raise e
            _error_num += 1

            return wrapper(*args, **kwargs)

    return wrapper


# 获取当前时间
def get_current_date():
    return time.strftime('%Y%m%d', time.localtime(time.time()))


# 获取当前时间戳
def get_current_time():
    return time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))


# 读取yaml文件数据
def get_yaml_content(read_path) -> dict:
    try:
        with open(read_path, mode='r', encoding="utf-8") as rf:
            return yaml.safe_load(rf)
    except:
        err = traceback.format_exc()
        print(f"open yaml files failed! \n{err}")
        raise err


# 写入yaml文件数据
def write_yaml_content(write_path, content):
    try:
        with open(write_path, mode='r', encoding="utf-8") as wf:
            wf.write(content)
    except:
        err = traceback.format_exc()
        print(f"write yaml files failed! \n{err}")
        raise err


# 读取csv文件数据_字典格式1
def get_csv_contents(filename):
    file = open(file=filename, mode='r', encoding='utf-8-sig')
    read = csv.DictReader(file)
    return list(read)


# 读取csv文件数据_字典格式2
def get_csv_content(filename):
    f = open(file=filename, mode='r', encoding='utf-8-sig')
    reader = list(csv.reader(f))
    # [{'csid':1,'page':1},{}]
    dictList = []
    for i in range(1, len(reader)):
        dict = {}  # 一条数据的字典
        data = reader[i]  # 一条csv数据
        keys = reader[0]  # key数据
        for j in range(len(keys)):
            dict[keys[j]] = data[j]  # 组装一条字典

        dictList.append(dict)  # 添加到列表

    return dictList


#  切换输入法
def switch_input(devices, name):
    """

    :param devices: 手机设备号
    :param name: ui_pytest_demo_master or yosemite
    :return:
    """
    try:
        if name == 'yosemite':
            os.system(f'adb -s {devices} shell ime set com.netease.nie.yosemite/.ime.ImeService')

        elif name == 'ui_pytest_demo_master':
            os.system(f'adb -s {devices} shell ime set io.ui_pytest_demo_master.settings/.AppiumIME')

        else:
            print("输入错误")
    except Exception as e:
        print("切换输入法方法失败:", e)


# 获取指定文件绝对路径
def filter_file(search_file_dir, target_file_name):
    all_content = os.listdir(search_file_dir)
    for per_file in all_content:
        file_path = os.path.join(search_file_dir, per_file)
        if os.path.isdir(file_path):
            final_file_path = filter_file(file_path, target_file_name)
            # 匹配到返回匹配到的值 否则继续
            if final_file_path:
                return final_file_path
            else:
                continue

        file_name = os.path.split(file_path)[1]
        if file_name == target_file_name:
            return file_path


# 读取yaml文件内容并返回list类型
def load_yaml_file(data_file, module_name, test_class_name, test_method_name) -> dict:
    yaml_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'datas')
    yaml_file = filter_file(yaml_dir, data_file)
    try:
        with open(yaml_file, mode='r', encoding="utf-8") as rf:
            content = yaml.safe_load(rf)
        parameters = content[module_name][test_class_name][test_method_name]
        # if len(parameters) > 1:
        return parameters

    except:
        err = traceback.format_exc()
        print(f"open yaml files failed! \n{err}")
        logging.error(f"open yaml files failed! \n{err}")
        sys.exit()


def send_dingTalk(at_mobiles=["17798068615", ]):
    webhook = "https://oapi.dingtalk.com/robot/send?access_token=b4c54f4298191a91b32be9941f14410f0bdabec4ef3ed7f20245004333804ec9"
    secret = "SEC817c602fb46721d499b396b3f19a5b39226646fd6a2e339f937ac86ca461d234"

    # 初始化机器人小丁
    xiaoding = DingtalkChatbot(webhook, secret, pc_slide=False, fail_notice=True)  # 方式一：通常初始化方式

    # Link消息(无法@指定用户)
    # xiaoding.send_link(title=title, text=text, message_url=message_url, pic_url=pic_url)

    # Markdown消息@指定用户
    xiaoding.send_markdown(title="接口自动化测试报告", text='接口自动化测试报告，请查收！\n'
                                                   '> ![美景](http://www.sinaimg.cn/dy/slidenews/5_img/2013_28/453_28488_469248.jpg)\n'
                                                   '> ###### [allure报告链接地址](http://www.thinkpage.cn/) \n',
                           at_mobiles=at_mobiles)


"""
user="2470162517@qq.com", password="pszxqzwiryvedihj", host="smtp.qq.com",
                 contents=["接口自动化测试报告，请查收！", "脚本：run.py", "作者:Project", "allure报告链接地址为:",
                           "https://service.mail.qq.com/cgi-bin/help?subtype=1&&id=20022&&no=1000729"],
                 to=['982152000@qq.com'], subject="接口自动化测试报告", attachments=None
"""


# 发送邮箱
def send_mail(user=None, password=None, host=None, contents=None, to=None, subject=None, attachments=None):
    """

    :param user:        用来发送邮件的邮箱
    :param password:    邮箱授权码
    :param host:        邮箱的smtp服务器地址
    :param contents:    邮箱正文，自定义        例：["Aritest测试报告", "脚本：test.air", "作者:AritestProject"] --> qq邮箱页面展示内容为换行展示（一个元素对应展示一行）
    :param to:          接收邮件的邮箱          例：多个收件人写list格式['123@qq.com','1255@qq.com']，单个收件人是str格式'1255@qq.com'
    :param subject:     邮件标题
    :param attachments: 附件地址，绝对路径       例：多个附件写list格式[r'd://log.txt', r'd://baidu_img.jpg',r"D:\AA\log.zip"]，单个附件是str格式r'd://baidu_img.jpg'  * 附件格式包括：txt，jpg，zip等等...
    :return:
    """
    user = user
    password = password
    host = host
    contents = contents
    to = to
    subject = subject
    attachments = None
    try:
        # 连接邮箱服务器 【由于通过yagmail发送文件，文件名会乱码，设置SMTP格式为即可：encoding="GBK"】
        yag = yagmail.SMTP(user=user, password=password, host=host, encoding="GBK")

        # 邮箱正文，自定义
        contents = contents

        # 发送邮箱
        yag.send(to=to, subject=subject, contents=contents, attachments=attachments, )
    except  Exception as e:
        print("发送邮箱方法失败:", e)


# 目录压缩zip格式方法
# :param dirpath: 需要导出的文件夹路径
# :param outFullName: 导出的zip压缩包的路径(含压缩包名称，此压缩文件为绝对路径)；【例 ： r"D:\AA\log.zip"】
def zipDir(dirpath, outFullName):
    try:
        zip = zipfile.ZipFile(outFullName, "w", zipfile.ZIP_DEFLATED)

        for path, dirnames, filenames in os.walk(dirpath):

            fpath = path.replace(dirpath, "")

            for filename in filenames:
                zip.write(os.path.join(path, filename), os.path.join(fpath, filename))

        zip.close()
    except Exception as e:
        print("目录压缩zip格式方法失败:", e)


# 读取写入ini文件
class iniUtils:

    def __init__(self, filename):
        self.filename = filename
        self.conf = ConfigParser()
        self.conf.read(filename, encoding="utf-8")

    # 读取ini文件
    def get_ini_content(self, section, option):
        """
        :param section:
        :param option:
        :return:
        """
        result = self.conf.get(section, option)
        return result

    # 写入ini文件 - 追加
    def write_ini_content(self, section, option, name):
        """
        写入ini文件时，如果本身有option值，就会进行写入覆盖，如果没有就会进行写入追加（支持中英文写入格式）
        :param section:
        :param option:
        :return:
        """

        # 设置_写入追加/覆盖section里面的option值
        self.conf.set(section, option, name)  # 写入中文
        # 设置之后必须设置文件格式:r追加/覆盖（非中文），r+追加/覆盖（中文）
        self.conf.write(open(self.filename, "r+", encoding="utf-8"))  # r+模式


class LinkMysqlDatabase:
    conf = ConfigParser()

    conf.read(MYSQL_PATH, encoding="utf-8")

    def __init__(self):

        host = self.conf.get('mysql', 'host')  # 数据库的ip地址
        user = self.conf.get('mysql', 'user')  # 数据库的账号
        password = self.conf.get('mysql', 'password')  # 数据库的密码
        port = self.conf.getint('mysql', 'port')  # mysql数据库的端口号
        self.usehost = self.conf.get('mysql', 'usehost')  # 选择某一个数据库进行查询
        self.mysql = pymysql.connect(host=host, user=user, password=password, port=port, charset="utf8",
                                     cursorclass=pymysql.cursors.DictCursor)
        self.cursor = self.mysql.cursor()

    def get_count(self, sql):
        """

        :param sql:  sql语句，例："select * from member where mobile_phone='13927427491';"
        :return:
        """
        try:
            # 进入所有数据库中的某一个数据库
            self.cursor.execute(self.usehost)
            # 输入sql语句
            count = self.cursor.execute(sql)
            return count
        except Exception as e:
            print("数据库查询共有多少条数据方法失败:", e)
        self.close()

    def get_query(self, sql):
        """

        :param sql:  sql语句，例："select * from member where mobile_phone='13927427491';"
        :return:
        """
        try:
            # 进入所有数据库中的某一个数据库
            self.cursor.execute(self.usehost)
            # 输入sql语句
            self.cursor.execute(sql)
            fetchone = self.cursor.fetchone()
            return fetchone
        except Exception as e:
            print("数据库查询单条数据方法失败:", e)
        self.close()

    def get_querys(self, sql):
        """

        :param sql:  sql语句，例："select * from member where mobile_phone='13927427491';"
        :return:
        """
        try:
            # 进入所有数据库中的某一个数据库
            self.cursor.execute(self.usehost)
            # 输入sql语句
            self.cursor.execute(sql)
            fetchall = self.cursor.fetchall()
            return fetchall
        except Exception as e:
            print("数据库查询多条数据方法失败:", e)
        self.close()

    def get_query_size(self, sql, size=None):
        """

        :param sql: sql语句，例："select * from member where mobile_phone='13927427491';"
        :param size: 行数 数字int类型
        :return:
        """
        try:
            # 进入所有数据库中的某一个数据库
            self.cursor.execute(self.usehost)
            # 输入sql语句
            self.cursor.execute(sql)
            if size:
                fetchmany = self.cursor.fetchmany(size=size)
                return fetchmany
            else:
                return "请输入数据库查询指定行数：size=数字int类型！"
        except Exception as e:
            print("数据库查询指定行数数据方法失败:", e)
        self.close()

    # 数据库插入/更新/删除数据
    def set(self, sql):
        """

        :param sql:  sql语句，例："select * from member where mobile_phone='13927427491';"
        :return:
        """
        try:
            # 进入所有数据库中的某一个数据库
            self.cursor.execute(self.usehost)
            # 输入sql语句
            self.cursor.execute(sql)
            # 提交数据库执行
            self.mysql.commit()
        except Exception as e:
            # 如果发生错误则回滚
            print("数据库插入/更新/删除数据方法失败:", e)
            self.mysql.rollback()
        self.close()

    def close(self):
        self.cursor.close()
        self.mysql.close()


# 连接数据库
class LinkPgDatabase:
    """
    eg:with LinkDatabase(username, password, host, port, link_database_name) as db:
            db.query_content(sql)
            db.insert_table_content(sql)

            extension: use records's api
                        db.db.query()
                        ...
    """

    def __init__(self, username, password, host, port, link_database_name):
        self._username = username
        self._password = password
        self._host = host
        self._port = port
        self._query_database_name = link_database_name

    def __enter__(self):
        # db = records.Database('postgresql://username:password@host.cn:port/query_database_name')
        self.db = records.Database(
            f'postgresql://{self._username}:{self._password}@{self._host}:{self._port}/{self._query_database_name}')
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.db.close()

    # 查询数据库
    def query_content(self, sql, as_dict=False, export=False, export_type='json', dataset=False):
        # sql ='SELECT * from public.filters WHERE id=787'

        rows = self.db.query(sql)
        if as_dict is True:
            return rows.all(as_dict=True)

        if export is True:
            if export_type.lower() == 'json':
                return rows.export('json')
            elif export_type.lower() == 'yaml':
                return rows.export('yaml')
            elif export_type.lower() == 'csv':
                return rows.export('csv')
            elif export_type.lower() == 'xls' or export_type.lower() == 'xlsx':
                return rows.export('xls')
        if dataset is True:
            return rows.dataset

        return rows

    # 插入数据
    def insert_content(self, sql, bulk=False):
        self.db.query(sql)
        if bulk is True:
            self.db.bulk_query(sql)

    # 建表
    def create_table(self, sql):
        self.db.query(sql)


class xlsxUtils:
    def __init__(self, filename):
        # xlsx文件绝对路径名称
        self.filename = filename
        # 获取 工作簿对象
        self.workbook = openpyxl.load_workbook(fr"{filename}")

    # 获取所有工作表的表名
    def sheetnames(self):
        try:
            sheetnames = self.workbook.sheetnames
            return sheetnames
        except Exception as e:
            return "输入格式有误:", e

    # 根据表名/索引值获取工作表对象
    def worksheet(self, name=None, num=None):
        """

        :param name: 表名，str类型
        :param num: 索引值，int类型
        :return:
        """
        try:
            if name and isinstance(name, str):
                worksheet = self.workbook[name]
                return worksheet

            elif num == 0:
                shenames = self.sheetnames()
                worksheet = self.workbook[shenames[0]]
                return worksheet

            elif num and isinstance(num, int):
                shenames = self.sheetnames()
                worksheet = self.workbook[shenames[num]]
                return worksheet

            else:
                return "输入格式有误！"
        except Exception as e:
            return "输入格式有误:", e

    # 获取工作表的表名title
    def title(self, name=None, num=None):
        try:
            worksheet = self.worksheet(name=name, num=num)
            res_name = worksheet.title
            return res_name
        except Exception as e:
            return "输入格式有误:", e

    # 获取工作表的“行数”max_row
    def max_row(self, name=None, num=None):
        try:
            worksheet = self.worksheet(name=name, num=num)
            rows = worksheet.max_row
            return rows
        except Exception as e:
            return "输入格式有误:", e

    # 获取工作表的“列数”max_column
    def max_column(self, name=None, num=None):
        try:
            worksheet = self.worksheet(name=name, num=num)
            columns = worksheet.max_column
            return columns
        except Exception as e:
            return "输入格式有误:", e

    # 按行方式获取表中的所有数据(格式为：[['北京', ' 石家庄'], [1, 2]]) --> ["参数":"参数"] ，["参数值":"参数值"]
    def rows(self, name=None, num=None):
        try:
            res_list = []
            worksheet = self.worksheet(name=name, num=num)
            for row in worksheet.rows:
                list = []
                for cell in row:
                    list.append(cell.value)
                res_list.append(list)
            return res_list
        except Exception as e:
            return "输入格式有误:", e

    # 按列方式获取表中的所有数据(格式为:[['北京', 1], [' 石家庄', 2]]) --> ["参数":"参数值"]，注：只适用于俩行数据，不推荐使用
    def columns(self, name=None, num=None):
        try:
            res_list = []
            worksheet = self.worksheet(name=name, num=num)
            for col in worksheet.columns:
                list = []
                for cell in col:
                    list.append(cell.value)
                res_list.append(list)
            return res_list
        except Exception as e:
            return "输入格式有误:", e

    # 获取表中几行几列的数据，如何没有数据就会返回None【注意：在Excel表，行和列都是从1开始计数的!!】
    def cells(self, row, column, name=None, num=None):
        """

        :param row: 行，int类型
        :param column: 列，int类型
        :return:
        """
        try:
            res_list = []
            worksheet = self.worksheet(name=name, num=num)
            for i in range(1, row + 1):
                list = []
                for j in range(1, column + 1):
                    value = worksheet.cell(row=i, column=j).value
                    list.append(value)
                res_list.append(list)
            return res_list
        except Exception as e:
            return "输入格式有误:", e

    # 获取表中某一单元格的数据
    # 精确读取表格中的某一单元格
    def value(self, pos="A1", name=None, num=None):
        """

        :param pos: xlsx坐标值，例：A1
        :return:
        """
        try:
            worksheet = self.worksheet(name=name, num=num)
            result = worksheet[pos].value
            return result
        except Exception as e:
            return "输入格式有误:", e

    # 获取表中某一单元格的数据，并重新覆盖写入新值
    def writeValue(self, pos="A1", values=None, name=None, num=None):
        """

        :param pos: xlsx坐标值，例：A1
        :param values: 写入的值，str类型
        :return:
        """
        try:
            worksheet = self.worksheet(name=name, num=num)
            worksheet[pos].value = values
            # 一旦做了修改，就要保存，保存的时候，要保证没有其它程序在使用当前文件。否则会报Permission Error
            self.workbook.save(self.filename)
            return f"写入成功,写入的值为:{values}"
        except Exception as e:
            return "输入格式有误:", e

    # 将行方式获取xlsx表数据整合成列表嵌套字典类型
    def getListDictRowsXlsx(self, name=None, num=None):
        reader = self.rows(name=name, num=num)
        # [{'csid':1,'page':1},{}]
        dictList = []
        for i in range(1, len(reader)):
            dict = {}  # 一条数据的字典
            data = reader[i]  # 一条csv数据
            keys = reader[0]  # key数据
            for j in range(len(keys)):
                dict[keys[j]] = data[j]  # 组装一条字典
            dictList.append(dict)  # 添加到列表

        return dictList

    # 将获取表中几行几列的xlsx表数据整合成列表嵌套字典类型
    def getListDictCellsXlsx(self, row, column, name=None, num=None):
        reader = self.cells(row, column, name=name, num=num)
        # [{'csid':1,'page':1},{}]
        dictList = []
        for i in range(1, len(reader)):
            dict = {}  # 一条数据的字典
            data = reader[i]  # 一条csv数据
            keys = reader[0]  # key数据
            for j in range(len(keys)):
                dict[keys[j]] = data[j]  # 组装一条字典

            dictList.append(dict)  # 添加到列表

        return dictList


# op项目由api获取cookie
class GetOpCookies:

    def __init__(self):
        self.session = requests.Session()

    def _get_account_session(self, base_url):
        res = self.session.get(base_url + '/authorize')
        assert 200 == res.status_code
        # print(res.status_code)

    def _get_user_session(self, username, password, base_url):
        payload = {
            'identity': username,
            'password': password
        }
        res = self.session.post(base_url + '/accounts/login', data=payload)
        assert 200 == res.status_code
        # print(res.status_code)

    @staticmethod
    def _cookie_format(cookies: dict, base_url) -> list:
        cookies_list = []
        for key, value in cookies.items():
            add_cookie = {'name': f'{key}', 'value': f'{value}', 'httponly': False,
                          'domain': f'{base_url.split("//")[1]}', 'path': '/', 'secure': False}
            cookies_list.append(add_cookie)
        return cookies_list

    def _get_cookie_from_redirect(self, url, base_url):
        res = self.session.get(url, allow_redirects=False)
        cookie_text = res.headers.get('Set-Cookie')
        redirect_url = res.headers.get('Location')
        if redirect_url:
            if not redirect_url.startswith("http"):
                redirect_url = base_url + redirect_url
        if res.status_code != 200:
            if cookie_text:
                cookie_key, cookie_value = cookie_text.split(";")[0].split('=')
                domain = urlparse(base_url).netloc
                # 添加cookie
                c = RequestsCookieJar()
                c.set(cookie_key, cookie_value, path='/', domain=domain)
                self.session.cookies.update(c)

            self._get_cookie_from_redirect(redirect_url, base_url)
        return

    # uat环境获取cookie
    def get_op_uat_cookies(self, username, password, uat_base_url):
        """:return: cookie
              type:dict
        """
        self._get_account_session(uat_base_url)
        self._get_user_session(username, password, uat_base_url)
        cookies = self.session.cookies
        return self._cookie_format(cookies, uat_base_url)

    # qa环境获取cookie
    def get_op_qa_cookie(self, username, password, qa_base_url):
        self._get_user_session(username, password, qa_base_url)
        self._get_cookie_from_redirect(qa_base_url + '/authorize', qa_base_url)
        cookies = dict(self.session.cookies)
        return self._cookie_format(cookies, qa_base_url)

